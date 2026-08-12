"""提报表解析器 — 支持 Excel / JSON / 文本三种输入方式."""

import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from geo_review.models import Submission
from geo_review.utils.text import clean_cell_value, is_example_row, split_list_field


# Excel 中文表头 → 字段键映射（兼容带 * 的必填标记）
_HEADER_MAP = {
    "任务名称": "task_name",
    "任务名称*": "task_name",
    "提交人": "submitter",
    "公司/品牌名称": "company_name",
    "公司/品牌名称*": "company_name",
    "产品/服务名称": "product_or_service",
    "产品/服务名称*": "product_or_service",
    "核心主题": "core_topic",
    "核心主题*": "core_topic",
    "核心信息点": "key_points",
    "核心信息点*": "key_points",
    "关键词": "keywords",
    "标准表述": "reference_copy",
    "允许写的事实/数据": "allowed_facts",
    "禁止出现的表述": "forbidden_claims",
    "禁止出现的表述*": "forbidden_claims",
    "禁止提及内容": "must_not_mention",
    "竞品名称": "competitor_names",
    "官网url": "official_urls",
    "官网url*": "official_urls",
    "官网url": "official_urls",
    "官网url*": "official_urls",
    "补充说明": "notes",
}

# 需要按分号拆分为列表的字段
_LIST_FIELDS = {
    "product_or_service",
    "key_points",
    "keywords",
    "reference_copy",
    "allowed_facts",
    "forbidden_claims",
    "must_not_mention",
    "competitor_names",
    "official_urls",
}

# 必填字段（业务层校验，与 schema 对齐）
_REQUIRED_FIELDS = {
    "task_name",
    "company_name",
    "product_or_service",
    "core_topic",
    "key_points",
    "forbidden_claims",
    "official_urls",
}


def _normalize_header(header: str) -> Optional[str]:
    """将 Excel 中文表头归一化为字段键."""
    if not header:
        return None
    key = str(header).strip()
    # 直接匹配
    if key in _HEADER_MAP:
        return _HEADER_MAP[key]
    # 忽略大小写再试一次
    low = key.lower()
    if low in _HEADER_MAP:
        return _HEADER_MAP[low]
    # 去除 * 号再匹配
    key_no_star = key.replace("*", "").strip()
    if key_no_star in _HEADER_MAP:
        return _HEADER_MAP[key_no_star]
    if key_no_star.lower() in _HEADER_MAP:
        return _HEADER_MAP[key_no_star.lower()]
    return None


def _row_to_dict(headers: List[str], row_values: List[Any]) -> Dict[str, Any]:
    """将 Excel 一行数据映射为字段字典."""
    result: Dict[str, Any] = {}
    for h, v in zip(headers, row_values):
        field_key = _normalize_header(h)
        if not field_key:
            continue
        cleaned = clean_cell_value(v)
        if cleaned is None:
            # 列表字段缺省为空列表，其他为 None
            if field_key in _LIST_FIELDS:
                result[field_key] = []
            else:
                result[field_key] = None
            continue
        if field_key in _LIST_FIELDS:
            result[field_key] = split_list_field(cleaned)
        else:
            result[field_key] = cleaned
    return result


def _validate_required(data: Dict[str, Any], source: str = "unknown") -> None:
    """校验必填字段是否齐全."""
    missing = [f for f in _REQUIRED_FIELDS if not data.get(f)]
    if missing:
        raise ValueError(
            f"[{source}] 缺少必填字段: {', '.join(missing)}"
        )


class SubmissionParser:
    """提报表解析器 — 统一入口.

    支持输入:
        - Excel 文件 (.xlsx / .xls)
        - JSON 文件 / 字符串 / 字典
        - 纯文本（未来扩展: NLP 抽取）

    返回:
        - Submission 对象（Pydantic 校验通过）
    """

    # ------------------------------------------------------------------
    # 公共入口
    # ------------------------------------------------------------------
    @classmethod
    def parse(
        cls,
        source: Union[str, Path, Dict[str, Any], bytes],
        *,
        format_hint: Optional[str] = None,
        sheet_name: Optional[str] = None,
        skip_example: bool = True,
    ) -> Submission:
        """自动识别并解析提报表.

        Args:
            source: 文件路径、JSON字符串、字典或文件二进制内容
            format_hint: 强制指定格式 ('excel' | 'json' | 'dict')
            sheet_name: Excel 工作表名，默认自动探测
            skip_example: 是否跳过疑似示例的行

        Returns:
            Submission 对象

        Raises:
            ValueError: 解析失败或校验不通过
        """
        fmt = (format_hint or "").lower()

        # 1) 字典直接构造
        if isinstance(source, dict):
            return cls.from_dict(source)

        # 2) 字符串: 先尝试 JSON，再视为路径
        if isinstance(source, str):
            stripped = source.strip()
            if stripped.startswith(("{", "[")):
                try:
                    return cls.from_json(stripped)
                except Exception:
                    pass
            # 视为文件路径
            return cls.from_file(
                Path(stripped), sheet_name=sheet_name, skip_example=skip_example
            )

        # 3) Path 对象
        if isinstance(source, Path):
            return cls.from_file(
                source, sheet_name=sheet_name, skip_example=skip_example
            )

        # 4) bytes — 按格式提示或扩展名解析
        if isinstance(source, bytes):
            if fmt in ("excel", "xlsx", "xls"):
                return cls._from_excel_bytes(source, sheet_name, skip_example)
            if fmt == "json":
                return cls.from_json(source.decode("utf-8"))
            if fmt in ("txt", "md"):
                text = source.decode("utf-8")
                stripped = text.strip()
                if stripped.startswith("{"):
                    return cls.from_json(text)
                return cls._from_text(text)
            raise ValueError(
                f"bytes 输入须显式指定 format_hint='excel'、'json' 或 'txt'，当前格式: {fmt}"
            )

        raise TypeError(f"不支持的输入类型: {type(source)}")

    # ------------------------------------------------------------------
    # 各格式专用方法
    # ------------------------------------------------------------------
    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> Submission:
        """从字典构造 Submission（Pydantic 自动校验）."""
        _validate_required(data, source="dict")
        return Submission.model_validate(data)

    @classmethod
    def from_json(cls, text: str) -> Submission:
        """从 JSON 字符串解析."""
        try:
            data = json.loads(text)
        except json.JSONDecodeError as exc:
            raise ValueError(f"JSON 解析失败: {exc}") from exc
        if not isinstance(data, dict):
            raise ValueError("JSON 顶层须为对象")
        return cls.from_dict(data)

    @classmethod
    def _from_text(cls, text: str) -> Submission:
        """解析文本格式的提报表（键值对格式）."""
        import re

        text = text.strip()

        if text.startswith("{"):
            return cls.from_json(text)

        data: Dict[str, Any] = {}
        lines = text.split("\n")
        for line in lines:
            line = line.strip()
            if not line:
                continue
            match = re.match(r"^(.+?)[:：]\s*(.+)$", line)
            if match:
                key = match.group(1).strip()
                value = match.group(2).strip()
                data[key] = value

        if not data:
            raise ValueError("文本提报表格式无效，请使用键值对格式如：任务名称：xxx")

        mapping = {
            "任务名称": "task_name",
            "公司名称": "company_name",
            "产品服务": "product_or_service",
            "核心主题": "core_topic",
            "要点": "key_points",
            "关键词": "keywords",
            "参考文案": "reference_copy",
            "允许事实": "allowed_facts",
            "禁用表述": "forbidden_claims",
            "禁止提及": "must_not_mention",
            "竞品名": "competitor_names",
            "官网地址": "official_urls",
            "备注": "notes",
        }

        result: Dict[str, Any] = {
            "product_or_service": ["未指定"],
            "core_topic": "未指定",
            "key_points": ["未指定"],
            "keywords": [],
            "reference_copy": [],
            "allowed_facts": [],
            "forbidden_claims": ["行业第一", "唯一", "100%", "最好", "最佳"],
            "must_not_mention": [],
            "competitor_names": [],
            "official_urls": ["https://example.com"],
            "notes": "无",
        }

        for cn_key, en_key in mapping.items():
            if cn_key in data:
                val = data[cn_key]
                if en_key in ["product_or_service", "key_points", "keywords",
                              "reference_copy", "allowed_facts", "forbidden_claims",
                              "must_not_mention", "competitor_names", "official_urls"]:
                    if isinstance(val, str):
                        val = [v.strip() for v in val.split("；") if v.strip()]
                    elif isinstance(val, list):
                        pass
                    else:
                        val = [str(val)]
                result[en_key] = val

        if "task_name" not in result or not result["task_name"]:
            raise ValueError("缺少必填字段: 任务名称")
        if "company_name" not in result or not result["company_name"]:
            raise ValueError("缺少必填字段: 公司名称")

        return Submission(**result)

    @classmethod
    def from_file(
        cls,
        path: Path,
        *,
        sheet_name: Optional[str] = None,
        skip_example: bool = True,
    ) -> Submission:
        """从文件自动识别格式并解析."""
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {path}")

        suffix = path.suffix.lower()

        if suffix in (".xlsx", ".xls"):
            return cls._from_excel_file(path, sheet_name, skip_example)

        if suffix == ".json":
            return cls.from_json(path.read_text(encoding="utf-8"))

        if fmt in ("txt", "md"):
            text = path.read_text(encoding="utf-8")
            stripped = text.strip()
            if stripped.startswith("{"):
                return cls.from_json(text)
            return cls._from_text(text)

        raise ValueError(f"不支持的文件格式: {suffix}")

    # ------------------------------------------------------------------
    # Excel 解析内部实现
    # ------------------------------------------------------------------
    @classmethod
    def _from_excel_file(
        cls,
        path: Path,
        sheet_name: Optional[str] = None,
        skip_example: bool = True,
    ) -> Submission:
        """从 Excel 文件解析提报表."""
        suffix = path.suffix.lower()

        if suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise ImportError("请安装 openpyxl: pip install openpyxl") from exc
            wb = load_workbook(path, data_only=True)
            return cls._parse_openpyxl(wb, sheet_name, skip_example)

        if suffix == ".xls":
            try:
                import xlrd
            except ImportError as exc:
                raise ImportError("请安装 xlrd: pip install xlrd") from exc
            book = xlrd.open_workbook(str(path))
            return cls._parse_xlrd(book, sheet_name, skip_example)

        raise ValueError(f"不支持的 Excel 格式: {suffix}")

    @classmethod
    def _from_excel_bytes(
        cls,
        data: bytes,
        sheet_name: Optional[str] = None,
        skip_example: bool = True,
    ) -> Submission:
        """从 Excel 二进制数据解析（根据文件头自动识别格式）."""
        if not data:
            raise ValueError("Excel 数据为空")

        # xlsx 文件头为 ZIP 格式 (50 4B 03 04)
        # xls 文件头为 OLE 格式 (D0 CF 11 E0)
        is_xlsx = data[:4] == b"PK\x03\x04"
        is_xls = data[:4] == b"\xd0\xcf\x11\xe0"

        if is_xlsx:
            try:
                from openpyxl import load_workbook
                from io import BytesIO
                wb = load_workbook(BytesIO(data), data_only=True)
                return cls._parse_openpyxl(wb, sheet_name, skip_example)
            except Exception as exc:
                raise ValueError(f"xlsx 解析失败: {exc}") from exc

        if is_xls:
            try:
                import xlrd
                book = xlrd.open_workbook(file_contents=data)
                return cls._parse_xlrd(book, sheet_name, skip_example)
            except Exception as exc:
                raise ValueError(f"xls 解析失败: {exc}") from exc

        raise ValueError("无法识别的 Excel 格式（非 xlsx/xls）")

    @classmethod
    def _parse_openpyxl(
        cls,
        workbook,
        sheet_name: Optional[str] = None,
        skip_example: bool = True,
    ) -> Submission:
        """使用 openpyxl 解析 .xlsx 文件."""
        ws = cls._get_sheet(workbook, sheet_name)

        # 读取表头（第1行）
        headers = [clean_cell_value(cell.value) for cell in ws[1]]
        if not any(headers):
            raise ValueError("Excel 第1行为空，无法识别表头")

        # 从第2行起遍历数据
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_data = _row_to_dict(headers, row)
            if skip_example and is_example_row(row_data):
                continue
            # 简单校验: 若关键必填字段缺失则跳过空行
            if not row_data.get("task_name"):
                continue
            _validate_required(row_data, source=f"Excel 第{row_idx}行")
            return Submission.model_validate(row_data)

        return Submission()

    @classmethod
    def _parse_xlrd(
        cls,
        book,
        sheet_name: Optional[str] = None,
        skip_example: bool = True,
    ) -> Submission:
        """使用 xlrd 解析 .xls 文件."""
        sheet = cls._get_xlrd_sheet(book, sheet_name)

        # 读取表头
        headers = [clean_cell_value(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
        if not any(headers):
            raise ValueError("Excel 第1行为空，无法识别表头")

        # 从第2行起遍历
        for row_idx in range(1, sheet.nrows):
            row_values = [sheet.cell_value(row_idx, c) for c in range(sheet.ncols)]
            row_data = _row_to_dict(headers, row_values)
            if skip_example and is_example_row(row_data):
                continue
            if not row_data.get("task_name"):
                continue
            _validate_required(row_data, source=f"Excel 第{row_idx + 1}行")
            return Submission.model_validate(row_data)

        return Submission()

    # ------------------------------------------------------------------
    # Sheet 选择辅助
    # ------------------------------------------------------------------
    @staticmethod
    def _get_sheet(workbook, sheet_name: Optional[str] = None):
        """openpyxl 工作表选择."""
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"工作表 '{sheet_name}' 不存在，可用: {workbook.sheetnames}"
                )
            return workbook[sheet_name]
        # 自动选择：优先 "提报信息"，否则取第一个可见 sheet
        for name in workbook.sheetnames:
            if "提报" in name or "信息" in name:
                return workbook[name]
        return workbook.active

    @staticmethod
    def _get_xlrd_sheet(book, sheet_name: Optional[str] = None):
        """xlrd 工作表选择."""
        if sheet_name:
            try:
                return book.sheet_by_name(sheet_name)
            except xlrd.XLRDError as exc:
                raise ValueError(
                    f"工作表 '{sheet_name}' 不存在，可用: {book.sheet_names()}"
                ) from exc
        for name in book.sheet_names():
            if "提报" in name or "信息" in name:
                return book.sheet_by_name(name)
        return book.sheet_by_index(0)

    # ------------------------------------------------------------------
    # 批量解析
    # ------------------------------------------------------------------
    @classmethod
    def parse_all_from_excel(
        cls,
        source: Union[Path, bytes],
        *,
        sheet_name: Optional[str] = None,
        skip_example: bool = True,
    ) -> List[Submission]:
        """从 Excel 批量解析所有任务行.

        Args:
            source: 文件路径(Path) 或 文件二进制内容(bytes)
            sheet_name: 指定工作表名
            skip_example: 是否跳过疑似示例行

        Returns:
            Submission 对象列表
        """
        submissions: List[Submission] = []

        # 统一获取 workbook / book 对象和解析参数
        if isinstance(source, bytes):
            if not source:
                raise ValueError("Excel 数据为空")
            is_xlsx = source[:4] == b"PK\x03\x04"
            is_xls = source[:4] == b"\xd0\xcf\x11\xe0"
            if is_xlsx:
                from openpyxl import load_workbook
                from io import BytesIO
                wb = load_workbook(BytesIO(source), data_only=True)
                sheets = wb.sheetnames
                ws = cls._get_sheet(wb, sheet_name)
                headers = [clean_cell_value(cell.value) for cell in ws[1]]
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                for row_idx, row in enumerate(rows, start=2):
                    row_data = _row_to_dict(headers, row)
                    if skip_example and is_example_row(row_data):
                        continue
                    if not row_data.get("task_name"):
                        continue
                    try:
                        _validate_required(row_data, source=f"Excel 第{row_idx}行")
                        submissions.append(Submission.model_validate(row_data))
                    except ValueError as exc:
                        print(f"[警告] 跳过第{row_idx}行: {exc}")
                return submissions
            if is_xls:
                import xlrd
                book = xlrd.open_workbook(file_contents=source)
                sheet = cls._get_xlrd_sheet(book, sheet_name)
                headers = [clean_cell_value(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
                for row_idx in range(1, sheet.nrows):
                    row_values = [sheet.cell_value(row_idx, c) for c in range(sheet.ncols)]
                    row_data = _row_to_dict(headers, row_values)
                    if skip_example and is_example_row(row_data):
                        continue
                    if not row_data.get("task_name"):
                        continue
                    try:
                        _validate_required(row_data, source=f"Excel 第{row_idx + 1}行")
                        submissions.append(Submission.model_validate(row_data))
                    except ValueError as exc:
                        print(f"[警告] 跳过第{row_idx + 1}行: {exc}")
                return submissions
            raise ValueError("无法识别的 Excel 格式（非 xlsx/xls）")

        # Path 对象处理
        path = Path(source)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {path}")
        suffix = path.suffix.lower()

        if suffix == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            ws = cls._get_sheet(wb, sheet_name)
            headers = [clean_cell_value(cell.value) for cell in ws[1]]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_data = _row_to_dict(headers, row)
                if skip_example and is_example_row(row_data):
                    continue
                if not row_data.get("task_name"):
                    continue
                try:
                    _validate_required(row_data, source=f"Excel 第{row_idx}行")
                    submissions.append(Submission.model_validate(row_data))
                except ValueError as exc:
                    print(f"[警告] 跳过第{row_idx}行: {exc}")

        elif suffix == ".xls":
            import xlrd
            book = xlrd.open_workbook(str(path))
            sheet = cls._get_xlrd_sheet(book, sheet_name)
            headers = [clean_cell_value(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
            for row_idx in range(1, sheet.nrows):
                row_values = [sheet.cell_value(row_idx, c) for c in range(sheet.ncols)]
                row_data = _row_to_dict(headers, row_values)
                if skip_example and is_example_row(row_data):
                    continue
                if not row_data.get("task_name"):
                    continue
                try:
                    _validate_required(row_data, source=f"Excel 第{row_idx + 1}行")
                    submissions.append(Submission.model_validate(row_data))
                except ValueError as exc:
                    print(f"[警告] 跳过第{row_idx + 1}行: {exc}")
        else:
            raise ValueError(f"不支持的 Excel 格式: {suffix}")

        return submissions
